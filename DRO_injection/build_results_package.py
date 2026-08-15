"""Assemble dist/A_within_one_year/ - the shareable results package for
the 9-spacecraft single-launch DRO phasing study.

Creates:
  dist/A_within_one_year/
    README.md
    Phase9_Results.xlsx        all numbers: per-spacecraft, scenario, comparison
    figures/                   every figure + animation produced so far
    data/                      per-spacecraft results CSV, daily fleet states CSV
    scripts/gmat/              runnable GMAT scripts (targeting x9, slots x9, replay)
    scripts/matlab/            plot_results.m - regenerates the plots from data/
    scripts/python/            the generators (optimisation, figures, GMAT builders)

Re-run any time; it rebuilds everything from the current outputs.
"""
import os, csv, shutil
from datetime import date, timedelta
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import sys

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))

from paths import (
    DIST_ROOT, GENERATED_GMAT, GMAT_OUTPUT, PYTHON_OUTPUT, ensure_runtime_dirs
)
from campaign_config import LAUNCH

ensure_runtime_dirs()
RES = DIST_ROOT / "A_within_one_year"
GMAT_OUT = GMAT_OUTPUT
PY_OUT = PYTHON_OUTPUT
NSAT  = 9
BUDGET = 2700.0
NOTES = {1: "long-way transfer (outer arc)", 2: "long-way transfer (outer arc)"}

for sub in ["figures", "data", "scripts/gmat", "scripts/matlab", "scripts/python"]:
    (RES / sub).mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------- gather numbers
# Python-model numbers from the campaign departure design (real-vE asymptote)
import json
with (PY_OUT / "campaign_departure.json").open("r", encoding="utf-8") as f:
    DEP = json.load(f)
py = {m["slot"]: dict(day=m["day"], dv1=m["dv1"] * 1000, dv2=m["dv2"] * 1000,
                      tot=m["total"] * 1000) for m in DEP["members"]}

gm, slot = {}, {}
for k in range(NSAT):
    r = (GMAT_OUT / f"g02_sc{k}_summary.txt").read_text().split()
    gm[k] = dict(dv1=float(r[0]) * 1000, dv2=float(r[1]) * 1000,
                 res=float(np.sqrt(float(r[5])**2 + float(r[6])**2 + float(r[7])**2)))
    r = (GMAT_OUT / f"g01_slot{k}.txt").read_text().split()
    slot[k] = dict(xE=float(r[2]) / 1e6, yE=float(r[3]) / 1e6, rsun=float(r[8]) / 1.495978707e8)

# ---------------------------------------------------------------- spreadsheet
wb = openpyxl.Workbook()
BOLD = Font(bold=True)
HDR  = PatternFill("solid", fgColor="D9E2F3")
RED  = PatternFill("solid", fgColor="F8CBAD")
GRN  = PatternFill("solid", fgColor="C6EFCE")

ws = wb.active
ws.title = "Per-spacecraft results"
head = ["S/C", "Slot (apse offset, d)", "Insertion day", "Insertion date",
        "Burn 1 Python (m/s)", "Burn 2 Python (m/s)", "Total Python (m/s)",
        "Burn 1 GMAT (m/s)", "Burn 2 GMAT (m/s)", "Total GMAT (m/s)",
        "GMAT - Python (m/s)", "GMAT targeting residual (km)",
        "Insertion xE (Mkm)", "Insertion yE (Mkm)", "Sun distance at insertion (au)",
        "Budget (m/s)", "Margin vs budget (m/s)", "Notes"]
ws.append(head)
for c in ws[1]:
    c.font, c.fill = BOLD, HDR
for k in range(NSAT):
    totg = gm[k]["dv1"] + gm[k]["dv2"]
    ws.append([f"S{k+1}", round(k * 365.25636 / 9, 2), int(py[k]["day"]),
               (LAUNCH + timedelta(days=py[k]["day"])).isoformat(),
               round(py[k]["dv1"]), round(py[k]["dv2"]), round(py[k]["tot"]),
               round(gm[k]["dv1"]), round(gm[k]["dv2"]), round(totg),
               round(totg - py[k]["tot"]), round(gm[k]["res"]),
               round(slot[k]["xE"], 2), round(slot[k]["yE"], 2), round(slot[k]["rsun"], 4),
               int(BUDGET), round(BUDGET - totg), NOTES.get(k, "")])
    ws.cell(ws.max_row, 17).fill = RED
worst = max(gm[k]["dv1"] + gm[k]["dv2"] for k in range(NSAT))
ws.append([])
ws.append(["VERDICT", "", "", "", "", "", "", "", "",
           f"worst {worst:.0f} m/s", "", "", "", "", "", "",
           "does NOT close", "all-inserted-within-365-d is infeasible"])
ws.cell(ws.max_row, 1).font = BOLD
for col, w in zip("ABCDEFGHIJKLMNOPQR", [6, 18, 13, 12, 18, 18, 17, 16, 16, 15, 16, 24, 14, 14, 24, 12, 19, 34]):
    ws.column_dimensions[col].width = w

ws = wb.create_sheet("Scenario & assumptions")
rows = [
    ("SCENARIO", ""),
    ("Constellation", "9 spacecraft, evenly phased (slots 40.58 d apart in apse epoch)"),
    ("Launch", "Single Ariane 64, direct escape, all 9 released on one hyperbola"),
    ("Launch epoch", f"{LAUNCH} 00:00 UTC - best 2045 epoch from scan_launch_dates.py; a DRO has no real launch window (full-year delta-v spread only ~314 m/s)"),
    ("Shared asymptote", "C3 = 4.0 km2/s2, in-plane angle 269 deg (optimised to minimise the worst member)"),
    ("Deployment requirement studied", "All 9 inserted within 365 days of launch (FD-105.2 is TBD in the RTM)"),
    ("Burns per spacecraft", "Burn 1 = departure trim just after separation; burn 2 = DRO insertion (anywhere on its slot's track)"),
    ("", ""),
    ("TARGET ORBIT", ""),
    ("DRO", "JPL Sun-Earth CR3BP catalogue ID 365 (planar)"),
    ("Period", "365.023318 days"),
    ("Max sunward amplitude", "22.122321 Mkm"),
    ("Max +/-y extent", "43.935877 Mkm"),
    ("ISR dwell per spacecraft", "34.52 days per revolution (9.46% of period)"),
    ("ISR distance margin", "5,879 km over the 21.5 Mkm requirement (razor-thin - drives FD-102.1/102.4)"),
    ("", ""),
    ("MODELS", ""),
    ("Python scan", "Planar heliocentric 2-body, Earth circular; validated to 0.73% (flux) vs GMAT/FD previously"),
    ("GMAT validation", "Full ephemeris DE405: Sun+Earth+Moon+Venus+Mars+Jupiter point masses + SRP; per-slot DROs re-converged; PD78 integrator"),
    ("Agreement", "Per-spacecraft totals within +/-300 m/s (<= 6%)"),
    ("Burn model", "Impulsive; finite-burn losses with the 10 N thruster NOT included (hours-long burns - flagged)"),
    ("", ""),
    ("KEY FINDINGS", ""),
    ("Verdict", f"Worst member {worst:.0f} m/s vs 2,700 m/s total propulsion budget - INFEASIBLE"),
    ("Why", "3 of 9 slots have no cheap perpendicular-crossing arrival within 365 d; timing offsets of 40-160 Mkm along-track in under a year inherently cost km/s-class delta-v"),
    ("FD-101.1", "The RTM's 'dv_insertion = 14.369 m/s' is actually the old GMAT script's TOTAL launchpad-to-DRO delta-v in km/s; real insertion is ~0.4-0.7 km/s"),
    ("Recommendation", "Staggered escape from a bound Earth loiter orbit: ~1.0-1.2 km/s per spacecraft, insertions day ~250-575 (see Architecture comparison)"),
]
for a, b in rows:
    ws.append([a, b])
    if b == "" and a:
        ws.cell(ws.max_row, 1).font = BOLD
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 130
for row in ws.iter_rows():
    row[1].alignment = Alignment(wrap_text=False)

ws = wb.create_sheet("Architecture comparison")
head = ["Architecture", "Per-s/c delta-v (worst)", "Fleet complete", "Closes 2700 m/s budget?", "Notes"]
ws.append(head)
for c in ws[1]:
    c.font, c.fill = BOLD, HDR
comp = [
    ["A. Single launch, direct escape, all inserted < 1 yr (THIS STUDY)", "4,260-5,730 m/s",
     "day 266 (~9 months)", "NO (needs > 2x budget)", "GMAT-verified; the 'why not' exhibit"],
    ["B. Single launch, bound Earth loiter, staggered escapes (RECOMMENDED)", "~1,000-1,250 m/s",
     "day ~575 (~19 months)", "YES (~1.5 km/s reserve)",
     "Every member flies the same cheap transfer; escape top-up ~560 m/s + insertion ~500-660 m/s; revives Earth-eclipse cases (AgZn battery basis); needs loiter-orbit design"],
    ["C. Single launch, direct escape, multi-rev drift orbits", "~2,100-2,600 m/s",
     "4-5 years", "MARGINAL", "Only if loiter is rejected; long deployment, thermal envelope of drift orbits differs"],
]
for r in comp:
    ws.append(r)
    ws.cell(ws.max_row, 4).fill = GRN if r[3].startswith("YES") else RED
for col, w in zip("ABCDE", [62, 22, 20, 24, 100]):
    ws.column_dimensions[col].width = w

# ---- ephemeris sheets (daily GMAT fleet states, launch -> day 996) ----
AUKM = 149597870.691
S0   = 1361.0     # W/m2 at 1 au; TCS_MODEL carries S0 = 1370 - scale by 1370/1361 if needed
daily = np.loadtxt(GMAT_OUT / "g03_all9_daily.txt")
daily[:, 0] -= daily[0, 0]

ws = wb.create_sheet("Ephemeris rotating (daily)")
head = ["Day after launch"] + sum([[f"S{k+1} xE (Mkm)", f"S{k+1} yE (Mkm)"] for k in range(NSAT)], [])
ws.append(["Earth-centred Sun-Earth rotating frame; sunward is negative x. "
           "GMAT full-ephemeris replay, launch 01 Jun 2044 -> day 996 (end of mission)."])
ws.append(head)
for c in ws[2]:
    c.font, c.fill = BOLD, HDR
for i in range(daily.shape[0]):
    row = [round(daily[i, 0], 2)]
    for k in range(NSAT):
        row += [round(daily[i, 1 + 3 * k] / 1e6, 4), round(daily[i, 2 + 3 * k] / 1e6, 4)]
    ws.append(row)
ws.freeze_panes = "B3"
ws.column_dimensions["A"].width = 16

ws = wb.create_sheet("Sun distance & flux (daily)")
ws.append([f"Sun distance (au) and solar flux (W/m2, S0 = {S0:.0f} at 1 au) per spacecraft. "
           "Transfer + DRO phases; use for thermal analysis. Note TCS_MODEL carries S0 = 1370."])
ws.append(["", "MIN r (au)"] + [""] * 8 + ["MAX flux (W/m2)"])
ws.append(["Summary per s/c:"] +
          [round(min(daily[:, 3 + 3 * k]) / AUKM, 4) for k in range(NSAT)] +
          [round(S0 / (min(daily[:, 3 + 3 * k]) / AUKM) ** 2) for k in range(NSAT)])
head = ["Day after launch"] + [f"S{k+1} r (au)" for k in range(NSAT)] + \
       [f"S{k+1} flux (W/m2)" for k in range(NSAT)]
ws.append(head)
for c in ws[4]:
    c.font, c.fill = BOLD, HDR
for i in range(daily.shape[0]):
    r_au = [daily[i, 3 + 3 * k] / AUKM for k in range(NSAT)]
    ws.append([round(daily[i, 0], 2)] + [round(r, 5) for r in r_au] +
              [round(S0 / r ** 2, 1) for r in r_au])
ws.freeze_panes = "B5"
ws.column_dimensions["A"].width = 16

ws = wb.create_sheet("Data dictionary")
for a, b in [
    ("data/per_spacecraft_results.csv", "The per-spacecraft table in CSV form"),
    ("data/fleet_daily.csv", "Daily states launch->day 996 from the GMAT replay: day, then per S1..S9 rotating-frame xE, yE (km) and Sun distance (km)"),
    ("figures/", "All figures + the deployment animation (phase9_anim.gif)"),
    ("scripts/gmat/g01_slot<k>.script", "Converges slot k's DRO (differential corrector) and reports its insertion state"),
    ("scripts/gmat/g02_sc<k>.script", "Targets spacecraft k's transfer + insertion; open in GMAT GUI to watch one spacecraft"),
    ("scripts/gmat/g03_all9_replay.script", "All 9 together, launch to end of mission - the GUI showpiece"),
    ("scripts/matlab/plot_results.m", "Regenerates trajectory / Sun-distance / delta-v plots from data/"),
    ("scripts/python/", "The optimisation + figure + GMAT-generator sources"),
]:
    ws.append([a, b])
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 120

try:
    wb.save(os.path.join(RES, "Phase9_Results.xlsx"))
except PermissionError:
    alt = os.path.join(RES, "Phase9_Results_NEW.xlsx")
    wb.save(alt)
    print("WARNING: Phase9_Results.xlsx is open in Excel and could not be replaced.")
    print(f"         Saved to {os.path.basename(alt)} - close Excel and rename it.")

# ---------------------------------------------------------------- data CSVs
with open(os.path.join(RES, "data", "per_spacecraft_results.csv"), "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["sc", "slot_apse_offset_d", "insertion_day", "insertion_date",
                "dv1_python_ms", "dv2_python_ms", "total_python_ms",
                "dv1_gmat_ms", "dv2_gmat_ms", "total_gmat_ms",
                "residual_km", "xE_Mkm", "yE_Mkm", "sun_au", "margin_vs_2700_ms", "notes"])
    for k in range(NSAT):
        totg = gm[k]["dv1"] + gm[k]["dv2"]
        w.writerow([f"S{k+1}", round(k * 365.25636 / 9, 2), int(py[k]["day"]),
                    (LAUNCH + timedelta(days=py[k]["day"])).isoformat(),
                    round(py[k]["dv1"]), round(py[k]["dv2"]), round(py[k]["tot"]),
                    round(gm[k]["dv1"]), round(gm[k]["dv2"]), round(totg),
                    round(gm[k]["res"]), round(slot[k]["xE"], 2), round(slot[k]["yE"], 2),
                    round(slot[k]["rsun"], 4), round(BUDGET - totg), NOTES.get(k, "")])

hdr = ["day"] + sum([[f"S{k+1}_xE_km", f"S{k+1}_yE_km", f"S{k+1}_rsun_km"] for k in range(NSAT)], [])
np.savetxt(os.path.join(RES, "data", "fleet_daily.csv"), daily,
           delimiter=",", header=",".join(hdr), comments="", fmt="%.6f")

# ---------------------------------------------------------------- copy figures + scripts
for source in PY_OUT.iterdir():
    if source.suffix.lower() in {".png", ".gif"}:
        shutil.copy2(source, RES / "figures" / source.name)

for source in GENERATED_GMAT.glob("*.script"):
    shutil.copy2(source, RES / "scripts" / "gmat" / source.name)

for source in (HERE / "gmat").glob("*.py"):
    shutil.copy2(source, RES / "scripts" / "python" / source.name)

for fn in [
    "campaign_config.py", "paths.py", "phase9_single_launch.py", "departure.py",
    "phase9_figures.py", "phase9_snapshot.py", "build_results_package.py",
]:
    shutil.copy2(HERE / fn, RES / "scripts" / "python" / fn)

# ---------------------------------------------------------------- MATLAB plotter
open(os.path.join(RES, "scripts", "matlab", "plot_results.m"), "w").write(r"""%% plot_results.m
% Regenerates the phasing-study plots from the packaged data.
% Run from this folder (scripts/matlab/). Figures saved to ../../figures/matlab/.
%
% Data: ../../data/fleet_daily.csv          daily GMAT fleet states, launch -> day 996
%       ../../data/per_spacecraft_results.csv   per-spacecraft delta-v table

clear; clc; close all;

dataDir = fullfile('..','..','data');
figDir  = fullfile('..','..','figures','matlab');
if ~exist(figDir,'dir'); mkdir(figDir); end

T  = readtable(fullfile(dataDir,'per_spacecraft_results.csv'));
D  = readmatrix(fullfile(dataDir,'fleet_daily.csv'));
day = D(:,1);
nSat = 9;
AU = 149597870.691;
cols = turbo(nSat);

%% 1. rotating-frame trajectories, launch -> end of mission
figure('Position',[60 60 850 950]); hold on;
for k = 1:nSat
    x = D(:,2+3*(k-1))/1e6;          % xE, Mkm
    y = D(:,3+3*(k-1))/1e6;          % yE, Mkm
    plot(x,y,'Color',cols(k,:),'LineWidth',1.0, ...
         'DisplayName',sprintf('S%d (%d m/s)',k,T.total_gmat_ms(k)));
    iIns = find(day >= T.insertion_day(k),1);
    plot(x(iIns),y(iIns),'s','Color',cols(k,:),'MarkerFaceColor',cols(k,:), ...
         'MarkerSize',7,'HandleVisibility','off');
end
plot(0,0,'o','MarkerFaceColor',[0.2 0.45 0.9],'MarkerEdgeColor','k', ...
     'MarkerSize',9,'DisplayName','Earth');
axis equal; grid on; box on;
xlabel('rotating x_E (Mkm, sunward negative)');
ylabel('rotating y_E (Mkm)');
title({'9 spacecraft, one launch: transfers + DRO to end of mission', ...
       'GMAT full-ephemeris replay (squares = insertion)'});
legend('Location','eastoutside');
exportgraphics(gcf, fullfile(figDir,'trajectories_rotating.png'), 'Resolution',200);

%% 2. Sun distance vs time
figure('Position',[80 80 1100 550]); hold on;
for k = 1:nSat
    r = D(:,4+3*(k-1))/AU;
    plot(day, r,'Color',cols(k,:),'LineWidth',1.0,'DisplayName',sprintf('S%d',k));
    iIns = find(day >= T.insertion_day(k),1);
    plot(day(iIns), r(iIns),'s','Color',cols(k,:),'MarkerFaceColor',cols(k,:), ...
         'MarkerSize',6,'HandleVisibility','off');
end
yline(1-22.122321e6/AU,':','DRO perihelion','HandleVisibility','off');
yline(1+22.122321e6/AU,':','DRO aphelion','HandleVisibility','off');
grid on; box on;
xlabel('days after launch (01 Jun 2044)');
ylabel('Sun distance (au)');
title('Sun distance per spacecraft, launch to end of mission (squares = insertion)');
legend('Location','eastoutside');
exportgraphics(gcf, fullfile(figDir,'sun_distance.png'), 'Resolution',200);

%% 3. delta-v per spacecraft: Python scan vs GMAT
figure('Position',[100 100 900 480]);
bar(categorical(T.sc), [T.total_python_ms T.total_gmat_ms]);
hold on;
yline(2700,'--r','2700 m/s propulsion budget','LineWidth',1.5);
grid on; box on;
ylabel('total \Deltav per spacecraft (m/s)');
legend({'Python 2-body scan','GMAT full ephemeris'},'Location','northwest');
title('All-inserted-within-365-days deployment: \Deltav per spacecraft');
exportgraphics(gcf, fullfile(figDir,'delta_v_comparison.png'), 'Resolution',200);

fprintf('\nPer-spacecraft results (GMAT-verified):\n');
disp(T(:,{'sc','insertion_day','dv1_gmat_ms','dv2_gmat_ms','total_gmat_ms','margin_vs_2700_ms'}));
fprintf('Figures written to %s\n', figDir);
""")

# ---------------------------------------------------------------- README
open(os.path.join(RES, "README.md"), "w").write(f"""# RESULTS - 9-spacecraft single-launch DRO phasing study

Built by `scripts/python/build_results_package.py` (re-run it to rebuild).

**Question:** can one Ariane 64 deliver 9 spacecraft into the JPL Sun-Earth DRO ID 365,
evenly phased, all inserted within one year of launch?

**Answer: no.** Best case (GMAT full-ephemeris verified) the worst spacecraft needs
**{worst:.0f} m/s** against the 2,700 m/s propulsion budget. Three of the nine phase slots
have no cheap insertion opportunity inside 365 days - this is geometry, not burn design.
**Recommended alternative:** staggered escape from a bound Earth loiter -
~1,000-1,250 m/s per spacecraft, fleet complete ~19 months after launch.

Everything is in `Phase9_Results.xlsx` (per-spacecraft table, scenario, architecture
comparison). `figures/` has all plots + the deployment animation. `scripts/gmat/
g03_all9_replay.script` opens in the GMAT GUI for the live visual. `scripts/matlab/
plot_results.m` regenerates plots from `data/`. See the Data dictionary sheet for details.
""")

print("Results package built at", RES)
print(f"worst member (GMAT): {worst:.0f} m/s vs {BUDGET:.0f} budget")
